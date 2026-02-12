import zipfile
import xml.etree.ElementTree as ET
from openpyxl import Workbook
import os

# Rutas
base_path = r"c:\Users\gperez\OneDrive - MKTO CATAL IMPORTACIONES, S.L\TALLER\SOLICITUDES DE DATOS\REVISION INDORME CHRISTIAN\INVENTARIO"

# Archivos a procesar
archivos = [
    "INVENTARIO TINTAS PUBLINDAL.xlsm",
    "INVENTARIO PLASTICO-CARTON.xlsm"
]

# Namespaces de Excel
ns = {
    'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
}

def extraer_datos(source_file, output_file):
    print(f"\n{'='*60}")
    print(f"Procesando: {os.path.basename(source_file)}")
    print(f"{'='*60}")
    
    try:
        with zipfile.ZipFile(source_file, 'r') as zip_ref:
            # Leer las cadenas compartidas
            shared_strings = []
            try:
                with zip_ref.open('xl/sharedStrings.xml') as f:
                    tree = ET.parse(f)
                    root = tree.getroot()
                    for si in root.findall('.//main:si', ns):
                        text_parts = []
                        for t in si.iter():
                            if t.text:
                                text_parts.append(t.text)
                        shared_strings.append(''.join(text_parts))
            except:
                pass
            
            print(f"Cadenas de texto: {len(shared_strings)}")
            
            # Leer nombres de hojas
            sheet_names = []
            with zip_ref.open('xl/workbook.xml') as f:
                tree = ET.parse(f)
                root = tree.getroot()
                for sheet in root.findall('.//main:sheet', ns):
                    sheet_names.append(sheet.get('name'))
            
            print(f"Hojas encontradas: {len(sheet_names)}")
            
            # Crear nuevo libro
            wb = Workbook()
            wb.remove(wb.active)
            
            # Procesar cada hoja
            for i, sheet_name in enumerate(sheet_names, 1):
                try:
                    with zip_ref.open(f'xl/worksheets/sheet{i}.xml') as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        
                        ws = wb.create_sheet(title=sheet_name[:31])
                        
                        for row in root.findall('.//main:row', ns):
                            for cell in row.findall('main:c', ns):
                                cell_ref = cell.get('r')
                                cell_type = cell.get('t')
                                
                                value_elem = cell.find('main:v', ns)
                                if value_elem is not None and value_elem.text:
                                    if cell_type == 's':
                                        try:
                                            idx = int(value_elem.text)
                                            value = shared_strings[idx] if idx < len(shared_strings) else value_elem.text
                                        except:
                                            value = value_elem.text
                                    else:
                                        try:
                                            if '.' in value_elem.text:
                                                value = float(value_elem.text)
                                            else:
                                                value = int(value_elem.text)
                                        except:
                                            value = value_elem.text
                                    
                                    ws[cell_ref] = value
                        
                        print(f"  ✓ {sheet_name}")
                        
                except Exception as e:
                    print(f"  ✗ {sheet_name}: {e}")
                    ws = wb.create_sheet(title=sheet_name[:31])
            
            # Guardar
            wb.save(output_file)
            print(f"\n✓ Guardado: {os.path.basename(output_file)}")
            return True
            
    except Exception as e:
        print(f"✗ Error: {e}")
        return False

# Procesar todos los archivos
for archivo in archivos:
    source = os.path.join(base_path, archivo)
    output = os.path.join(base_path, archivo.replace('.xlsm', '_RECUPERADO.xlsx'))
    
    if os.path.exists(source):
        extraer_datos(source, output)
    else:
        print(f"Archivo no encontrado: {archivo}")

print(f"\n{'='*60}")
print("¡PROCESO COMPLETADO!")
print(f"{'='*60}")
