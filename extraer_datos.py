import zipfile
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import re

# Rutas
base_path = r"c:\Users\gperez\OneDrive - MKTO CATAL IMPORTACIONES, S.L\TALLER\SOLICITUDES DE DATOS\REVISION INDORME CHRISTIAN\INVENTARIO"
source_file = os.path.join(base_path, "INVENTARIO TINTAS PUBLINDAL.xlsm")
output_file = os.path.join(base_path, "INVENTARIO_DATOS_RECUPERADOS.xlsx")

# Namespaces de Excel
ns = {
    'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
}

print("Extrayendo datos del archivo Excel protegido...")

# Abrir el archivo xlsm como ZIP
with zipfile.ZipFile(source_file, 'r') as zip_ref:
    # Leer las cadenas compartidas (sharedStrings.xml)
    shared_strings = []
    try:
        with zip_ref.open('xl/sharedStrings.xml') as f:
            tree = ET.parse(f)
            root = tree.getroot()
            for si in root.findall('.//main:si', ns):
                # Obtener todo el texto del elemento si
                text_parts = []
                for t in si.iter():
                    if t.text:
                        text_parts.append(t.text)
                shared_strings.append(''.join(text_parts))
    except Exception as e:
        print(f"Nota: No se encontraron sharedStrings: {e}")
    
    print(f"Cadenas de texto encontradas: {len(shared_strings)}")
    
    # Leer los nombres de las hojas desde workbook.xml
    sheet_names = []
    with zip_ref.open('xl/workbook.xml') as f:
        tree = ET.parse(f)
        root = tree.getroot()
        for sheet in root.findall('.//main:sheet', ns):
            sheet_names.append(sheet.get('name'))
    
    print(f"Hojas encontradas: {len(sheet_names)}")
    for name in sheet_names:
        print(f"  - {name}")
    
    # Crear nuevo libro de Excel
    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto
    
    # Procesar cada hoja
    for i, sheet_name in enumerate(sheet_names, 1):
        print(f"\nProcesando hoja {i}/{len(sheet_names)}: {sheet_name}")
        
        try:
            with zip_ref.open(f'xl/worksheets/sheet{i}.xml') as f:
                tree = ET.parse(f)
                root = tree.getroot()
                
                # Crear hoja en el nuevo libro
                ws = wb.create_sheet(title=sheet_name[:31])  # Excel limita a 31 caracteres
                
                # Encontrar todas las filas
                rows_data = {}
                for row in root.findall('.//main:row', ns):
                    row_num = int(row.get('r'))
                    
                    for cell in row.findall('main:c', ns):
                        cell_ref = cell.get('r')
                        cell_type = cell.get('t')  # 's' = shared string, 'n' = number, etc.
                        
                        # Obtener valor
                        value_elem = cell.find('main:v', ns)
                        if value_elem is not None and value_elem.text:
                            if cell_type == 's':
                                # Es un índice a shared strings
                                try:
                                    idx = int(value_elem.text)
                                    value = shared_strings[idx] if idx < len(shared_strings) else value_elem.text
                                except:
                                    value = value_elem.text
                            else:
                                # Intentar convertir a número
                                try:
                                    if '.' in value_elem.text:
                                        value = float(value_elem.text)
                                    else:
                                        value = int(value_elem.text)
                                except:
                                    value = value_elem.text
                            
                            # Escribir en la celda
                            ws[cell_ref] = value
                
                print(f"  ✓ Hoja '{sheet_name}' procesada")
                
        except Exception as e:
            print(f"  ✗ Error procesando hoja {sheet_name}: {e}")
            # Crear hoja vacía
            ws = wb.create_sheet(title=sheet_name[:31])
    
    # Guardar el nuevo archivo
    print(f"\nGuardando archivo: {output_file}")
    wb.save(output_file)
    print("✓ Archivo guardado exitosamente!")
    print(f"\nAbre el archivo: INVENTARIO_DATOS_RECUPERADOS.xlsx")
